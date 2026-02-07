"""
Imports Router - Session 15
Handles import validation, preview, and bulk operations
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from datetime import datetime, date as date_type
from typing import List, Optional
import pandas as pd
import io

from ..database import get_db
from ..models import Batch, Container, AgeGroup, Supplier, Variety, Greenhouse
from ..auth import get_current_user
from pydantic import BaseModel

router = APIRouter()  # No prefix here - main.py adds /api/imports


# ============================================
# SCHEMAS
# ============================================

class ValidationError(BaseModel):
    row: int
    column: str
    value: str
    error: str

class ImportPreviewResponse(BaseModel):
    total_rows: int
    valid_rows: int
    error_rows: int
    errors: List[ValidationError]
    preview_data: List[dict]
    can_import: bool

class BulkUpdateRequest(BaseModel):
    batch_ids: List[int]
    field: str
    value: str

class BulkAssignRequest(BaseModel):
    batch_ids: List[int]
    greenhouse_id: int

class BulkDeleteRequest(BaseModel):
    batch_ids: List[int]
    confirm: bool = False

# ============================================
# HELPER: Auto-generate container number
# ============================================

def generate_container_number(db: Session, arrival_date: date_type) -> str:
    """Generate container number: CNT-YYYYMMDD-NNN"""
    date_str = arrival_date.strftime("%Y%m%d")
    prefix = f"CNT-{date_str}-"
    
    existing = db.query(Container).filter(
        Container.container_number.like(f"{prefix}%")
    ).all()
    
    if existing:
        max_num = 0
        for c in existing:
            try:
                num = int(c.container_number.split("-")[-1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
        next_num = max_num + 1
    else:
        next_num = 1
    
    return f"{prefix}{next_num:03d}"


def find_or_create_container(
    db: Session, 
    arrival_date: date_type,
    container_cache: dict
) -> int:
    """
    Find existing container or create new one.
    Groups by arrival_date ONLY (1 container can have multiple suppliers).
    """
    cache_key = arrival_date.isoformat()
    
    if cache_key in container_cache:
        return container_cache[cache_key]
    
    existing = db.query(Container).filter(
        Container.arrival_date == arrival_date
    ).first()
    
    if existing:
        container_cache[cache_key] = existing.id
        return existing.id
    
    container_number = generate_container_number(db, arrival_date)
    
    new_container = Container(
        container_number=container_number,
        arrival_date=arrival_date,
        status="active",
        notes="Auto-created from batch import"
    )
    
    db.add(new_container)
    db.flush()
    
    container_cache[cache_key] = new_container.id
    return new_container.id
def find_or_create_age_group(
    db: Session,
    container_id: int,
    age_group_name: str,
    age_group_cache: dict
) -> int:
    """
    Find existing age group or create new one.
    Groups by container_id + age_group_name.
    """
    cache_key = (container_id, age_group_name.strip().lower())
    
    if cache_key in age_group_cache:
        return age_group_cache[cache_key]
    
    existing = db.query(AgeGroup).filter(
        and_(
            AgeGroup.container_id == container_id,
            AgeGroup.name == age_group_name.strip()
        )
    ).first()
    
    if existing:
        age_group_cache[cache_key] = existing.id
        return existing.id
    
    new_age_group = AgeGroup(
        container_id=container_id,
        name=age_group_name.strip(),
        status="active"
    )
    
    db.add(new_age_group)
    db.flush()
    
    age_group_cache[cache_key] = new_age_group.id
    return new_age_group.id


# ============================================
# IMPORT VALIDATION & PREVIEW
# ============================================

@router.post("/batches/preview", response_model=ImportPreviewResponse)
async def preview_batch_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Preview and validate batch import file before actual import.
    Returns validation errors with row numbers for easy fixing.
    """
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx, .xls) or CSV")
    
    try:
        contents = await file.read()
        
        # Read file based on extension
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Required columns mapping
        required_columns = {
            'batch_code': ['batch_code', 'code', 'kode_batch', 'kode', 'batch_number'],
            'supplier': ['supplier', 'supplier_name', 'nama_supplier'],
            'variety': ['variety', 'variety_name', 'varietas', 'nama_varietas'],
            'quantity': ['quantity', 'qty', 'jumlah', 'initial_quantity'],
            'arrival_date': ['arrival_date', 'date', 'tanggal', 'tanggal_datang', 'date_received'],
            'age_group': ['age_group', 'kelompok_umur', 'stok_bulan', 'age']
        }
        
        # Map columns
        column_mapping = {}
        for required, alternatives in required_columns.items():
            for alt in alternatives:
                if alt in df.columns:
                    column_mapping[required] = alt
                    break
        
        # Check missing required columns
        missing_cols = [col for col in required_columns.keys() if col not in column_mapping]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_cols)}. "
                       f"Expected columns: batch_code, supplier, variety, quantity, arrival_date"
            )
        
        # Get existing data for validation
        existing_batch_codes = {b.batch_number for b in db.query(Batch.batch_number).all()}
        suppliers = {s.supplier_name.lower(): s.id for s in db.query(Supplier).all()}
        varieties = {v.variety_name.lower(): v.id for v in db.query(Variety).all()}
        
        errors = []
        valid_count = 0
        error_row_indices = set()
        preview_data = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (header is row 1)
            row_errors = []
            row_data = {
                'row_number': row_num,
                'batch_code': str(row.get(column_mapping.get('batch_code'), '')).strip(),
                'supplier': str(row.get(column_mapping.get('supplier'), '')).strip(),
                'variety': str(row.get(column_mapping.get('variety'), '')).strip(),
                'quantity': row.get(column_mapping.get('quantity')),
                'arrival_date': row.get(column_mapping.get('arrival_date')),
                'age_group': str(row.get(column_mapping.get('age_group'), '')).strip() if column_mapping.get('age_group') else '',
                'status': 'valid',
                'errors': []
            }
            
            # Validate batch_code
            batch_code = row_data['batch_code']
            if not batch_code or batch_code == 'nan':
                row_errors.append(ValidationError(
                    row=row_num, column='batch_code', value=batch_code, error='Batch code is required'
                ))
            elif batch_code in existing_batch_codes:
                row_errors.append(ValidationError(
                    row=row_num, column='batch_code', value=batch_code, error='Batch code already exists'
                ))
            
            # Validate supplier
            supplier = row_data['supplier'].lower()
            if not supplier or supplier == 'nan':
                row_errors.append(ValidationError(
                    row=row_num, column='supplier', value=row_data['supplier'], error='Supplier is required'
                ))
            elif supplier not in suppliers:
                row_errors.append(ValidationError(
                    row=row_num, column='supplier', value=row_data['supplier'], 
                    error=f'Supplier not found. Available: {", ".join(list(suppliers.keys())[:5])}...'
                ))
            
            # Validate variety
            variety = row_data['variety'].lower()
            if not variety or variety == 'nan':
                row_errors.append(ValidationError(
                    row=row_num, column='variety', value=row_data['variety'], error='Variety is required'
                ))
            elif variety not in varieties:
                row_errors.append(ValidationError(
                    row=row_num, column='variety', value=row_data['variety'],
                    error=f'Variety not found. Available: {", ".join(list(varieties.keys())[:5])}...'
                ))
            
            # Validate quantity
            try:
                qty = int(row_data['quantity'])
                if qty <= 0:
                    row_errors.append(ValidationError(
                        row=row_num, column='quantity', value=str(qty), error='Quantity must be greater than 0'
                    ))
                row_data['quantity'] = qty
            except (ValueError, TypeError):
                row_errors.append(ValidationError(
                    row=row_num, column='quantity', value=str(row_data['quantity']), error='Quantity must be a number'
                ))
            
            # Validate arrival_date
            try:
                if pd.isna(row_data['arrival_date']):
                    row_errors.append(ValidationError(
                        row=row_num, column='arrival_date', value='', error='Arrival date is required'
                    ))
                else:
                    date_val = pd.to_datetime(row_data['arrival_date'])
                    row_data['arrival_date'] = date_val.strftime('%Y-%m-%d')
            except Exception:
                row_errors.append(ValidationError(
                    row=row_num, column='arrival_date', value=str(row_data['arrival_date']),
                    error='Invalid date format. Use YYYY-MM-DD'
                ))
            
            # Track errors
            if row_errors:
                error_row_indices.add(row_num)
                row_data['status'] = 'error'
                row_data['errors'] = [e.error for e in row_errors]
                errors.extend(row_errors)
            else:
                valid_count += 1
            
            preview_data.append(row_data)
        
        return ImportPreviewResponse(
            total_rows=len(df),
            valid_rows=valid_count,
            error_rows=len(error_row_indices),
            errors=errors,
            preview_data=preview_data[:100],  # Limit preview to 100 rows
            can_import=valid_count > 0
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")


@router.post("/batches/execute")
async def execute_batch_import(
    file: UploadFile = File(...),
    skip_errors: bool = True,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Execute batch import after preview.
    Set skip_errors=True to import valid rows only, False to abort on any error.
    """
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx, .xls) or CSV")
    
    try:
        contents = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
        
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Column mapping (same as preview)
        required_columns = {
            'batch_code': ['batch_code', 'code', 'kode_batch', 'kode', 'batch_number'],
            'supplier': ['supplier', 'supplier_name', 'nama_supplier'],
            'variety': ['variety', 'variety_name', 'varietas', 'nama_varietas'],
            'quantity': ['quantity', 'qty', 'jumlah', 'initial_quantity'],
            'arrival_date': ['arrival_date', 'date', 'tanggal', 'tanggal_datang', 'date_received'],
            'age_group': ['age_group', 'kelompok_umur', 'stok_bulan', 'age']
        }
        
        column_mapping = {}
        for required, alternatives in required_columns.items():
            for alt in alternatives:
                if alt in df.columns:
                    column_mapping[required] = alt
                    break
        
        # Get lookup data
        existing_batch_codes = {b.batch_number for b in db.query(Batch.batch_number).all()}
        suppliers = {s.supplier_name.lower(): s.id for s in db.query(Supplier).all()}
        varieties = {v.variety_name.lower(): v.id for v in db.query(Variety).all()}
        
        imported = 0
        skipped = 0
        errors = []
        containers_created = 0
        container_cache = {}
        age_groups_created = 0
        age_group_cache = {}
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            try:
                batch_code = str(row.get(column_mapping.get('batch_code'), '')).strip()
                supplier_name = str(row.get(column_mapping.get('supplier'), '')).strip().lower()
                variety_name = str(row.get(column_mapping.get('variety'), '')).strip().lower()
                quantity = int(row.get(column_mapping.get('quantity')))
                arrival_date = pd.to_datetime(row.get(column_mapping.get('arrival_date'))).date()
                
                # Skip if batch_code exists
                if batch_code in existing_batch_codes:
                    if skip_errors:
                        skipped += 1
                        errors.append(f"Row {row_num}: Batch code '{batch_code}' already exists")
                        continue
                    else:
                        raise HTTPException(status_code=400, detail=f"Row {row_num}: Batch code already exists")
                
                # Get supplier and variety IDs
                supplier_id = suppliers.get(supplier_name)
                variety_id = varieties.get(variety_name)
                
                if not supplier_id or not variety_id:
                    if skip_errors:
                        skipped += 1
                        errors.append(f"Row {row_num}: Invalid supplier or variety")
                        continue
                    else:
                        raise HTTPException(status_code=400, detail=f"Row {row_num}: Invalid supplier or variety")
                
                # Auto-create container (group by arrival_date)
                old_cache_size = len(container_cache)
                container_id = find_or_create_container(
                    db=db,
                    arrival_date=arrival_date,
                    container_cache=container_cache
                )
                if len(container_cache) > old_cache_size:
                    containers_created += 1
                
                # Auto-create age group if provided
                age_group_id = None
                age_group_name = str(row.get(column_mapping.get('age_group'), '')).strip() if column_mapping.get('age_group') else ''
                if age_group_name and age_group_name != 'nan':
                    old_ag_size = len(age_group_cache)
                    age_group_id = find_or_create_age_group(
                        db=db,
                        container_id=container_id,
                        age_group_name=age_group_name,
                        age_group_cache=age_group_cache
                    )
                    if len(age_group_cache) > old_ag_size:
                        age_groups_created += 1
                
                # Create batch WITH container_id and age_group_id
                new_batch = Batch(
                    batch_number=batch_code,
                    supplier_id=supplier_id,
                    variety_id=variety_id,
                    quantity_expected=quantity,
                    quantity_actual=quantity,
                    qty_a1_start=quantity,
                    qty_a1_current=quantity,
                    date_received=arrival_date,
                    date_a1_entry=arrival_date,
                    current_phase='A1',
                    is_completed=False,
                    container_id=container_id,
                    age_group_id=age_group_id
                )
                
                db.add(new_batch)
                existing_batch_codes.add(batch_code)  # Prevent duplicates within same import
                imported += 1
                
            except HTTPException:
                raise
            except Exception as e:
                if skip_errors:
                    skipped += 1
                    errors.append(f"Row {row_num}: {str(e)}")
                else:
                    db.rollback()
                    raise HTTPException(status_code=400, detail=f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(df),
            "containers_created": containers_created,
            "age_groups_created": age_groups_created,
            "errors": errors[:20]  # Return first 20 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ============================================
# FIX EXISTING BATCHES (one-time migration)
# ============================================

@router.post("/batches/fix-containers")
async def fix_existing_batch_containers(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """One-time fix: Create containers for batches without container_id."""
    
    user_role = current_user.get('role') if isinstance(current_user, dict) else getattr(current_user, 'role', None)
    if user_role != 'admin':
        raise HTTPException(status_code=403, detail="Only admin can run this fix")
    
    orphan_batches = db.query(Batch).filter(Batch.container_id == None).all()
    
    if not orphan_batches:
        return {"success": True, "message": "No orphan batches found.", "fixed": 0, "containers_created": 0}
    
    container_cache = {}
    fixed = 0
    containers_created = 0
    
    for batch in orphan_batches:
        if batch.date_received:
            old_cache_size = len(container_cache)
            container_id = find_or_create_container(
                db=db,
                arrival_date=batch.date_received,
                container_cache=container_cache
            )
            if len(container_cache) > old_cache_size:
                containers_created += 1
            batch.container_id = container_id
            fixed += 1
    
    db.commit()
    
    return {
        "success": True,
        "message": f"Fixed {fixed} batches, created {containers_created} containers",
        "fixed": fixed,
        "containers_created": containers_created
    }

# ============================================
# BULK OPERATIONS
# ============================================

@router.post("/batches/bulk-update")
async def bulk_update_batches(
    request: BulkUpdateRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Bulk update batch field (status, phase, notes)"""
    
    allowed_fields = ['current_phase', 'notes', 'is_completed']
    
    if request.field not in allowed_fields:
        raise HTTPException(
            status_code=400, 
            detail=f"Field '{request.field}' not allowed for bulk update. Allowed: {allowed_fields}"
        )
    
    if request.field == 'current_phase' and request.value not in ['A1', 'A2', 'A3', 'A4', 'A5']:
        raise HTTPException(status_code=400, detail="Invalid phase. Must be A1-A5")
    
    if request.field == 'is_completed':
        value = request.value.lower() in ['true', '1', 'yes']
        updated = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).update(
            {request.field: value},
            synchronize_session=False
        )
    else:
        updated = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).update(
            {request.field: request.value},
            synchronize_session=False
        )
    
    db.commit()
    
    return {
        "success": True,
        "updated_count": updated,
        "field": request.field,
        "new_value": request.value
    }


@router.post("/batches/bulk-assign-greenhouse")
async def bulk_assign_greenhouse(
    request: BulkAssignRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Bulk assign batches to a greenhouse"""
    
    greenhouse = db.query(Greenhouse).filter(Greenhouse.id == request.greenhouse_id).first()
    if not greenhouse:
        raise HTTPException(status_code=404, detail="Greenhouse not found")
    
    # Check capacity
    current_plants = db.query(func.sum(
        Batch.qty_a1_current + Batch.qty_a2_current + Batch.qty_a3_current + Batch.qty_a4_current
    )).filter(
        and_(Batch.current_greenhouse_id == request.greenhouse_id, Batch.is_completed == False)
    ).scalar() or 0
    
    batches_to_assign = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).all()
    total_new_plants = sum(
        (b.qty_a1_current or 0) + (b.qty_a2_current or 0) + (b.qty_a3_current or 0) + (b.qty_a4_current or 0)
        for b in batches_to_assign
    )
    
    if greenhouse.capacity and current_plants + total_new_plants > greenhouse.capacity:
        raise HTTPException(
            status_code=400,
            detail=f"Capacity exceeded. Greenhouse capacity: {greenhouse.capacity}, "
                   f"Current: {current_plants}, Adding: {total_new_plants}"
        )
    
    updated = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).update(
        {"current_greenhouse_id": request.greenhouse_id},
        synchronize_session=False
    )
    
    db.commit()
    
    utilization = ((current_plants + total_new_plants) / greenhouse.capacity * 100) if greenhouse.capacity else 0
    
    return {
        "success": True,
        "updated_count": updated,
        "greenhouse": greenhouse.greenhouse_name,
        "new_utilization": f"{utilization:.1f}%"
    }


@router.post("/batches/bulk-delete")
async def bulk_delete_batches(
    request: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Bulk delete batches with confirmation"""
    
    # Check role
    user_role = current_user.get('role') if isinstance(current_user, dict) else getattr(current_user, 'role', None)
    if user_role not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="Only admin or supervisor can delete batches")
    
    if not request.confirm:
        # Return preview of what will be deleted
        batches = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).all()
        return {
            "confirm_required": True,
            "batches_to_delete": [
                {
                    "id": b.id,
                    "batch_number": b.batch_number,
                    "current_quantity": (b.qty_a1_current or 0) + (b.qty_a2_current or 0) + (b.qty_a3_current or 0) + (b.qty_a4_current or 0),
                    "phase": b.current_phase
                } for b in batches
            ],
            "total_plants_affected": sum(
                (b.qty_a1_current or 0) + (b.qty_a2_current or 0) + (b.qty_a3_current or 0) + (b.qty_a4_current or 0)
                for b in batches
            ),
            "message": "Set 'confirm': true to proceed with deletion"
        }
    
    # Perform deletion
    deleted = db.query(Batch).filter(Batch.id.in_(request.batch_ids)).delete(synchronize_session=False)
    db.commit()
    
    return {
        "success": True,
        "deleted_count": deleted
    }


# ============================================
# IMPORT TEMPLATE DOWNLOAD
# ============================================

@router.get("/templates/batches")
async def download_batch_import_template(
    current_user: dict = Depends(get_current_user)
):
    """Download Excel template for batch import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Import Template"
    
    # Headers
    headers = ["batch_code", "supplier", "variety", "quantity", "arrival_date", "notes"]
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Example data
    examples = [
        ["B-2024-001", "PT Anggrek Jaya", "Gan Lin Yukidian V3", 5000, "2024-01-15", "First batch"],
        ["B-2024-002", "CV Bunga Makmur", "Sogo Yukidian", 3000, "2024-01-20", ""],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["BATCH IMPORT INSTRUCTIONS"],
        [""],
        ["Required Columns:"],
        ["- batch_code: Unique batch identifier (e.g., B-2024-001)"],
        ["- supplier: Must match existing supplier name exactly"],
        ["- variety: Must match existing variety name exactly"],
        ["- quantity: Number of plants (positive integer)"],
        ["- arrival_date: Date in YYYY-MM-DD format"],
        [""],
        ["Optional Columns:"],
        ["- notes: Additional notes for the batch"],
        [""],
        ["Tips:"],
        ["- Remove example rows before importing"],
        ["- Ensure supplier and variety names match exactly"],
        ["- Use Preview function to validate before importing"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=row[0] if row else "")
    
    ws2.column_dimensions['A'].width = 60
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=batch_import_template.xlsx"}
    )
