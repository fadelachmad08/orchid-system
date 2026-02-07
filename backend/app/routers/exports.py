"""
Export Router - Session 15
Handles all export operations: Batches, Containers, Phase Transitions, Loss Reports
Supports Excel (.xlsx) and CSV (.csv) formats
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from datetime import datetime, date
from typing import Optional, List
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Batch, Container, PhaseTransition, Supplier, Variety, Greenhouse
from ..auth import get_current_user

router = APIRouter()  # No prefix here - main.py adds /api/exports

# ============================================
# STYLING HELPERS
# ============================================

def apply_header_style(cell):
    """Apply consistent header styling"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_cell_style(cell):
    """Apply consistent cell styling"""
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.alignment = Alignment(vertical="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(length + 2, 50)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


# ============================================
# EXPORT BATCHES
# ============================================

@router.get("/batches")
async def export_batches(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    phase: Optional[str] = Query(None, description="Filter by phase (A1-A5)"),
    supplier_id: Optional[int] = Query(None, description="Filter by supplier"),
    variety_id: Optional[int] = Query(None, description="Filter by variety"),
    greenhouse_id: Optional[int] = Query(None, description="Filter by greenhouse"),
    status: Optional[str] = Query(None, description="Filter by status"),
    date_from: Optional[date] = Query(None, description="Filter from date"),
    date_to: Optional[date] = Query(None, description="Filter to date"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export batches with advanced filters"""
    
    # Build query with filters
    query = db.query(Batch).outerjoin(Supplier, Batch.supplier_id == Supplier.id).outerjoin(Variety, Batch.variety_id == Variety.id)
    
    if phase:
        query = query.filter(Batch.current_phase == phase.upper())
    if supplier_id:
        query = query.filter(Batch.supplier_id == supplier_id)
    if variety_id:
        query = query.filter(Batch.variety_id == variety_id)
    if greenhouse_id:
        query = query.filter(Batch.current_greenhouse_id == greenhouse_id)
    if status:
        if status == 'active':
            query = query.filter(Batch.is_completed == False)
        elif status == 'completed':
            query = query.filter(Batch.is_completed == True)
    if date_from:
        query = query.filter(Batch.date_received >= date_from)
    if date_to:
        query = query.filter(Batch.date_received <= date_to)
    
    batches = query.all()
    
    # Prepare data
    headers = [
        "Batch Number", "Supplier", "Variety", "Current Phase",
        "Qty Expected", "Qty Received", "Qty Current", "Initial Loss",
        "Date Received", "Is Completed", "Notes"
    ]
    
    rows = []
    for batch in batches:
        rows.append([
            batch.batch_number,
            batch.supplier.supplier_name if batch.supplier else "-",
            batch.variety.variety_name if batch.variety else "-",
            batch.current_phase,
            batch.quantity_expected or 0,
            batch.quantity_actual or 0,
            (batch.qty_a1_current or 0) + (batch.qty_a2_current or 0) + (batch.qty_a3_current or 0) + (batch.qty_a4_current or 0),
            batch.initial_loss or 0,
            str(batch.date_received) if batch.date_received else "-",
            "Yes" if batch.is_completed else "No",
            batch.notes or "-"
        ])
    
    # Generate file
    if format == "csv":
        return generate_csv_response(headers, rows, "batches_export")
    else:
        return generate_excel_response(headers, rows, "Batches Export", "batches_export")


# ============================================
# EXPORT CONTAINERS
# ============================================

@router.get("/containers")
async def export_containers(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    supplier_id: Optional[int] = Query(None, description="Filter by supplier"),
    date_from: Optional[date] = Query(None, description="Filter from arrival date"),
    date_to: Optional[date] = Query(None, description="Filter to arrival date"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export containers with batch details"""
    
    query = db.query(Container).outerjoin(Supplier, Container.supplier_id == Supplier.id)
    
    if supplier_id:
        query = query.filter(Container.supplier_id == supplier_id)
    if date_from:
        query = query.filter(Container.arrival_date >= date_from)
    if date_to:
        query = query.filter(Container.arrival_date <= date_to)
    
    containers = query.all()
    
    headers = [
        "Container Number", "Supplier", "Arrival Date", 
        "Total Expected", "Total Received", "Initial Loss", "Notes"
    ]
    
    rows = []
    for c in containers:
        rows.append([
            c.container_number,
            c.supplier.supplier_name if c.supplier else "-",
            str(c.arrival_date) if c.arrival_date else "-",
            c.total_expected or 0,
            c.total_received or 0,
            c.initial_loss or 0,
            c.notes or "-"
        ])
    
    if format == "csv":
        return generate_csv_response(headers, rows, "containers_export")
    else:
        return generate_excel_response(headers, rows, "Containers Export", "containers_export")


# ============================================
# EXPORT PHASE TRANSITIONS
# ============================================

@router.get("/phase-transitions")
async def export_phase_transitions(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    batch_id: Optional[int] = Query(None),
    from_phase: Optional[str] = Query(None),
    to_phase: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export phase transition history"""
    
    query = db.query(PhaseTransition).join(Batch, PhaseTransition.batch_id == Batch.id)
    
    if batch_id:
        query = query.filter(PhaseTransition.batch_id == batch_id)
    if from_phase:
        query = query.filter(PhaseTransition.from_phase == from_phase.upper())
    if to_phase:
        query = query.filter(PhaseTransition.to_phase == to_phase.upper())
    if date_from:
        query = query.filter(PhaseTransition.transition_date >= date_from)
    if date_to:
        query = query.filter(PhaseTransition.transition_date <= date_to)
    
    transitions = query.order_by(PhaseTransition.transition_date.desc()).all()
    
    headers = [
        "Batch Number", "From Phase", "To Phase", 
        "Qty Before", "Qty Moved", "Qty Loss", "Loss Reason",
        "Transition Date", "Notes"
    ]
    
    rows = []
    for t in transitions:
        rows.append([
            t.batch.batch_number if t.batch else "-",
            t.from_phase,
            t.to_phase,
            t.quantity_before or 0,
            t.quantity_moved or 0,
            t.quantity_loss or 0,
            t.loss_reason or "-",
            str(t.transition_date) if t.transition_date else "-",
            t.notes or "-"
        ])
    
    if format == "csv":
        return generate_csv_response(headers, rows, "phase_transitions_export")
    else:
        return generate_excel_response(headers, rows, "Phase Transitions", "phase_transitions_export")


# ============================================
# EXPORT LOSS REPORT
# ============================================

@router.get("/loss-report")
async def export_loss_report(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    group_by: str = Query("phase", pattern="^(phase|supplier|variety)$"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export loss report grouped by phase, supplier, or variety"""
    
    query = db.query(PhaseTransition).join(Batch).filter(PhaseTransition.quantity_loss > 0)
    
    if date_from:
        query = query.filter(PhaseTransition.transition_date >= date_from)
    if date_to:
        query = query.filter(PhaseTransition.transition_date <= date_to)
    
    transitions = query.all()
    
    if group_by == "phase":
        headers = ["Phase Transition", "Total Loss Qty", "Avg Loss %", "Count"]
        grouped_data = {}
        
        for t in transitions:
            key = f"{t.from_phase} → {t.to_phase}"
            if key not in grouped_data:
                grouped_data[key] = {"total_loss": 0, "total_before": 0, "count": 0}
            grouped_data[key]["total_loss"] += (t.quantity_loss or 0)
            grouped_data[key]["total_before"] += (t.quantity_before or 0)
            grouped_data[key]["count"] += 1
        
        rows = []
        for key, data in grouped_data.items():
            avg_loss = (data["total_loss"] / data["total_before"] * 100) if data["total_before"] > 0 else 0
            rows.append([key, data["total_loss"], round(avg_loss, 2), data["count"]])
    
    elif group_by == "supplier":
        headers = ["Supplier", "Total Batches", "Total Loss Qty", "Avg Loss %"]
        grouped_data = {}
        
        for t in transitions:
            supplier = t.batch.supplier.supplier_name if t.batch and t.batch.supplier else "Unknown"
            if supplier not in grouped_data:
                grouped_data[supplier] = {"batches": set(), "total_loss": 0, "total_initial": 0}
            grouped_data[supplier]["batches"].add(t.batch_id)
            grouped_data[supplier]["total_loss"] += (t.quantity_loss or 0)
            if t.batch:
                grouped_data[supplier]["total_initial"] = max(
                    grouped_data[supplier]["total_initial"], 
                    t.batch.quantity_actual or 0
                )
        
        rows = []
        for supplier, data in grouped_data.items():
            avg_loss = (data["total_loss"] / data["total_initial"] * 100) if data["total_initial"] > 0 else 0
            rows.append([supplier, len(data["batches"]), data["total_loss"], round(avg_loss, 2)])
    
    else:  # variety
        headers = ["Variety", "Total Batches", "Total Loss Qty", "Avg Loss %"]
        grouped_data = {}
        
        for t in transitions:
            variety = t.batch.variety.variety_name if t.batch and t.batch.variety else "Unknown"
            if variety not in grouped_data:
                grouped_data[variety] = {"batches": set(), "total_loss": 0, "total_initial": 0}
            grouped_data[variety]["batches"].add(t.batch_id)
            grouped_data[variety]["total_loss"] += (t.quantity_loss or 0)
            if t.batch:
                grouped_data[variety]["total_initial"] = max(
                    grouped_data[variety]["total_initial"],
                    t.batch.quantity_actual or 0
                )
        
        rows = []
        for variety, data in grouped_data.items():
            avg_loss = (data["total_loss"] / data["total_initial"] * 100) if data["total_initial"] > 0 else 0
            rows.append([variety, len(data["batches"]), data["total_loss"], round(avg_loss, 2)])
    
    if format == "csv":
        return generate_csv_response(headers, rows, f"loss_report_by_{group_by}")
    else:
        return generate_excel_response(headers, rows, f"Loss Report by {group_by.title()}", f"loss_report_by_{group_by}")


# ============================================
# EXPORT INVENTORY SUMMARY
# ============================================

@router.get("/inventory-summary")
async def export_inventory_summary(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Export inventory summary by phase and greenhouse"""
    
    # By Phase
    phase_data = db.query(
        Batch.current_phase,
        func.count(Batch.id).label("batch_count"),
        func.sum(Batch.qty_a1_current + Batch.qty_a2_current + Batch.qty_a3_current + Batch.qty_a4_current).label("total_plants")
    ).filter(Batch.is_completed == False).group_by(Batch.current_phase).all()
    
    # By Greenhouse
    greenhouse_data = db.query(
        Greenhouse.greenhouse_name,
        Greenhouse.capacity,
        func.count(Batch.id).label("batch_count"),
        func.sum(Batch.qty_a1_current + Batch.qty_a2_current + Batch.qty_a3_current + Batch.qty_a4_current).label("total_plants")
    ).outerjoin(Batch, and_(Batch.current_greenhouse_id == Greenhouse.id, Batch.is_completed == False)
    ).group_by(Greenhouse.id).all()
    
    if format == "csv":
        # CSV: combine both summaries
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["=== INVENTORY BY PHASE ==="])
        writer.writerow(["Phase", "Batch Count", "Total Plants"])
        for row in phase_data:
            writer.writerow([row.current_phase, row.batch_count, row.total_plants or 0])
        
        writer.writerow([])
        writer.writerow(["=== INVENTORY BY GREENHOUSE ==="])
        writer.writerow(["Greenhouse", "Capacity", "Batch Count", "Total Plants", "Utilization %"])
        for row in greenhouse_data:
            utilization = (row.total_plants or 0) / row.capacity * 100 if row.capacity and row.capacity > 0 else 0
            writer.writerow([row.greenhouse_name, row.capacity or 0, row.batch_count, row.total_plants or 0, round(utilization, 1)])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=inventory_summary_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    else:
        # Excel: two sheets
        wb = Workbook()
        
        # Sheet 1: By Phase
        ws1 = wb.active
        ws1.title = "By Phase"
        headers1 = ["Phase", "Batch Count", "Total Plants"]
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        for row_idx, row in enumerate(phase_data, 2):
            ws1.cell(row=row_idx, column=1, value=row.current_phase)
            ws1.cell(row=row_idx, column=2, value=row.batch_count)
            ws1.cell(row=row_idx, column=3, value=row.total_plants or 0)
            for col in range(1, 4):
                apply_cell_style(ws1.cell(row=row_idx, column=col))
        
        auto_adjust_columns(ws1)
        
        # Sheet 2: By Greenhouse
        ws2 = wb.create_sheet("By Greenhouse")
        headers2 = ["Greenhouse", "Capacity", "Batch Count", "Total Plants", "Utilization %"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            apply_header_style(cell)
        
        for row_idx, row in enumerate(greenhouse_data, 2):
            utilization = (row.total_plants or 0) / row.capacity * 100 if row.capacity and row.capacity > 0 else 0
            ws2.cell(row=row_idx, column=1, value=row.greenhouse_name)
            ws2.cell(row=row_idx, column=2, value=row.capacity or 0)
            ws2.cell(row=row_idx, column=3, value=row.batch_count)
            ws2.cell(row=row_idx, column=4, value=row.total_plants or 0)
            ws2.cell(row=row_idx, column=5, value=round(utilization, 1))
            for col in range(1, 6):
                apply_cell_style(ws2.cell(row=row_idx, column=col))
        
        auto_adjust_columns(ws2)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventory_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )


# ============================================
# HELPER FUNCTIONS
# ============================================

def generate_csv_response(headers: List[str], rows: List[List], filename_prefix: str):
    """Generate CSV file response"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

def generate_excel_response(headers: List[str], rows: List[List], sheet_title: str, filename_prefix: str):
    """Generate Excel file response with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_style(cell)
    
    # Auto-adjust columns
    auto_adjust_columns(ws)
    
    # Add filter
    if rows:
        ws.auto_filter.ref = ws.dimensions
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
